import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from datetime import date
import re

# =========================
# 基础设置
# =========================

excel_file = "Specialty_Coffee_Market_Radar_3.1_RMB.xlsx"
sheet_name = "Raw_Data"
today = date.today().isoformat()
USD_TO_RMB = 7.2

headers = {"User-Agent": "Mozilla/5.0"}

# =========================
# 工具函数：写入Excel
# =========================

def write_row(ws, data):
    ws.insert_rows(2)

    ws.cell(row=2, column=1, value=today)
    ws.cell(row=2, column=2, value=data["roaster"])
    ws.cell(row=2, column=3, value=data["country"])
    ws.cell(row=2, column=4, value=data["name"])
    ws.cell(row=2, column=5, value=data["origin"])
    ws.cell(row=2, column=6, value=data["process"])
    ws.cell(row=2, column=7, value=data["size"])
    ws.cell(row=2, column=8, value=data["price"])
    ws.cell(row=2, column=10, value=data["stock"])
    ws.cell(row=2, column=11, value="Yes")
    ws.cell(row=2, column=12, value="=H2/G2")
    ws.cell(row=2, column=13, value=data["variety"])

# =========================
# 读取历史数据（用于去重+补货）
# =========================

def load_history(ws):
    history = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = row[3]
        stock = row[9]
        if name:
            history[name] = stock
    return history

# =========================
# Botz（JSON）
# =========================

def fetch_botz(ws, history):
    print("\n===== BOTZ =====")

    url = "https://botz-coffee.com/products.json"
    res = requests.get(url)
    products = res.json().get("products", [])

    count = 0

    for p in products:
        name = p["title"]

        variants = p["variants"]

        v = variants[0]

        price_rmb = round(float(v["price"]) * USD_TO_RMB, 2)
        stock = "Yes" if v["available"] else "No"

        if stock == "No":
            price_rmb = 0

        restock = ""
        if history.get(name) == "No" and stock == "Yes":
            restock = "Restock"

        write_row(ws, {
            "roaster": "Botz Coffee",
            "country": "USA",
            "name": name,
            "origin": "Unknown",
            "process": "Unknown",
            "size": 150,
            "price": price_rmb,
            "stock": stock,
            "variety": "Unknown"
        })

        print(name, stock, restock)
        count += 1

    print(f"Botz 完成：{count}")

# =========================
# Sey（JSON）
# =========================

def fetch_sey(ws, history):
    print("\n===== SEY =====")

    url = "https://www.seycoffee.com/products.json"
    res = requests.get(url)
    products = res.json().get("products", [])

    count = 0

    for p in products:
        name = p["title"]

        if "subscription" in name.lower():
            continue

        variants = p["variants"]

        price_rmb = None
        stock = "No"

        for v in variants:
            if "250g" in v["title"].lower():
                price_rmb = round(float(v["price"]) * USD_TO_RMB, 2)
                stock = "Yes" if v["available"] else "No"
                break

        if price_rmb is None and variants:
            v = variants[0]
            price_rmb = round(float(v["price"]) * USD_TO_RMB, 2)
            stock = "Yes" if v["available"] else "No"

        if stock == "No":
            price_rmb = 0

        restock = ""
        if history.get(name) == "No" and stock == "Yes":
            restock = "Restock"

        write_row(ws, {
            "roaster": "Sey Coffee",
            "country": "USA",
            "name": name,
            "origin": "Unknown",
            "process": "Unknown",
            "size": 250,
            "price": price_rmb,
            "stock": stock,
            "variety": "Unknown"
        })

        print(name, stock, restock)
        count += 1

    print(f"Sey 完成：{count}")

# =========================
# Hydrangea（HTML）
# =========================

def fetch_hydrangea(ws, history):
    print("\n===== HYDRANGEA =====")

    base_url = "https://hydrangea.coffee"

    res = requests.get(base_url, headers=headers)
    soup = BeautifulSoup(res.text, "html.parser")

    links = set()

    for a in soup.find_all("a", href=True):
        if "/products/" in a["href"]:
            links.add(base_url + a["href"])

    count = 0

    for link in links:

        try:
            res = requests.get(link, headers=headers)
            soup = BeautifulSoup(res.text, "html.parser")

            title_tag = soup.find("h1")
            if not title_tag:
                continue

            name = title_tag.get_text(strip=True)

            options = soup.find_all("option")

            price_rmb = None
            stock = "Yes"

            for opt in options:
                text = opt.get_text().lower()

                if "114g" in text or "4oz" in text:

                    if "sold" in text:
                        stock = "No"
                        price_rmb = 0
                        break

                    m = re.search(r"\$([0-9]+\.?[0-9]*)", text)
                    if m:
                        price_rmb = round(float(m.group(1)) * USD_TO_RMB, 2)
                        break

            if price_rmb is None:
                page_text = soup.get_text().lower()

                if "sold out" in page_text:
                    stock = "No"
                    price_rmb = 0

            restock = ""
            if history.get(name) == "No" and stock == "Yes":
                restock = "Restock"

            write_row(ws, {
                "roaster": "Hydrangea Coffee",
                "country": "USA",
                "name": name,
                "origin": "Unknown",
                "process": "Unknown",
                "size": 114,
                "price": price_rmb,
                "stock": stock,
                "variety": "Unknown"
            })

            print(name, stock, restock)
            count += 1

        except:
            continue

    print(f"Hydrangea 完成：{count}")

# =========================
# 总控函数
# =========================

def run_all():
    wb = load_workbook(excel_file)
    ws = wb[sheet_name]

    history = load_history(ws)

    fetch_botz(ws, history)
    fetch_sey(ws, history)
    fetch_hydrangea(ws, history)

    wb.save(excel_file)

    print("\n🎉 全部完成")

# =========================
# 运行
# =========================

if __name__ == "__main__":
    run_all()